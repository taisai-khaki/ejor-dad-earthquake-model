from __future__ import annotations

import hashlib
import json
import math
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from ejor_dad.checkpoint import atomic_write_dataframe, atomic_write_text


RAW = Path("data_work/noto/raw")
PREPARED = Path("data_work/noto/prepared")
CORE_ZONE_CODES = ["17202", "17204", "17205", "17461", "17463"]
CENTER_CODES = ["17201", *CORE_ZONE_CODES]
THROUGHPUT_PER_BED = 4.0

MUNICIPALITIES = {
    "17201": {"name_en": "Kanazawa", "name_ja": "金沢市", "full_destroyed": 31},
    "17202": {"name_en": "Nanao", "name_ja": "七尾市", "full_destroyed": 506},
    "17204": {"name_en": "Wajima", "name_ja": "輪島市", "full_destroyed": 2293},
    "17205": {"name_en": "Suzu", "name_ja": "珠洲市", "full_destroyed": 1738},
    "17461": {"name_en": "Anamizu", "name_ja": "穴水町", "full_destroyed": 395},
    "17463": {"name_en": "Noto", "name_ja": "能登町", "full_destroyed": 246},
}

CORRIDOR_OBSERVATIONS = [
    {
        "link_id": "corr_kanazawa_nanao",
        "label": "Kanazawa-Nanao",
        "tail_code": "17201",
        "head_code": "17202",
        "feature_tokens": ("金沢市", "七尾市"),
        "normal_minutes": 45.0,
        "disrupted_minutes": 50.0,
        "recovery_points": 0,
        "disrupted_observation": "2024-01-25 directional mean at 09:00",
    },
    {
        "link_id": "corr_nanao_anamizu",
        "label": "Nanao-Anamizu",
        "tail_code": "17202",
        "head_code": "17461",
        "feature_tokens": ("七尾市", "穴水町"),
        "normal_minutes": 30.0,
        "disrupted_minutes": 100.0,
        "recovery_points": 0,
        "disrupted_observation": "2024-01-11 MLIT intercity travel-time GeoJSON",
    },
    {
        "link_id": "corr_anamizu_wajima",
        "label": "Anamizu-Wajima",
        "tail_code": "17461",
        "head_code": "17204",
        "feature_tokens": ("穴水町", "輪島市"),
        "normal_minutes": 30.0,
        "disrupted_minutes": 80.0,
        "recovery_points": 7,
        "disrupted_observation": "2024-01-11 MLIT intercity travel-time GeoJSON",
    },
    {
        "link_id": "corr_anamizu_noto",
        "label": "Anamizu-Noto",
        "tail_code": "17461",
        "head_code": "17463",
        "feature_tokens": ("穴水町", "能登町"),
        "normal_minutes": 40.0,
        "disrupted_minutes": 70.0,
        "recovery_points": 0,
        "disrupted_observation": "2024-01-25 directional mean at 09:00",
    },
    {
        "link_id": "corr_anamizu_suzu",
        "label": "Anamizu-Suzu",
        "tail_code": "17461",
        "head_code": "17205",
        "feature_tokens": ("穴水町", "珠洲市"),
        "normal_minutes": 50.0,
        "disrupted_minutes": 90.0,
        "recovery_points": 3,
        "disrupted_observation": "2024-01-11 MLIT intercity travel-time GeoJSON",
    },
]

OPERATIONAL_SHARES = {
    "17201": 0.20,
    "17202": 0.55,
    "17204": 0.35,
    "17205": 0.35,
    "17461": 0.35,
    "17463": 0.35,
}


def main() -> None:
    PREPARED.mkdir(parents=True, exist_ok=True)
    population = load_population()
    hospitals = load_hospitals()
    coordinates, route_features = load_road_geometries()
    recovery_counts = load_recovery_counts()

    zones = build_zones(population)
    centers = build_centers(hospitals, coordinates)
    corridors = build_corridors(route_features, recovery_counts)
    snapshots = summarize_snapshots()
    sources = build_source_manifest()
    coverage = build_coverage_table()

    write_table(zones, "noto_zones")
    write_table(hospitals, "noto_hospitals")
    write_table(centers, "noto_centers")
    write_table(corridors, "noto_corridors")
    write_table(snapshots, "noto_road_snapshots")
    write_table(sources, "noto_source_manifest")
    write_table(coverage, "noto_data_coverage")

    payload = {
        "zone_count": len(zones),
        "center_count": len(centers),
        "corridor_count": len(corridors),
        "total_population": float(zones["population"].sum()),
        "baseline_at_risk_population": float(zones["at_risk_population"].sum()),
        "reported_hospital_beds": float(centers["reported_beds"].sum()),
        "modeled_existing_capacity": float(centers["existing_capacity"].sum()),
        "throughput_per_bed": THROUGHPUT_PER_BED,
        "calibration_locked_before_optimization": True,
    }
    atomic_write_text(PREPARED / "noto_preparation_summary.json", json.dumps(payload, indent=2, ensure_ascii=False))
    print(json.dumps(payload, indent=2, ensure_ascii=False))


def load_population() -> pd.DataFrame:
    path = RAW / "ishikawa_statistics" / "R4_population_machine_readable.xlsx"
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    worksheet = workbook["11"]
    rows: list[dict[str, Any]] = []
    for values in worksheet.iter_rows(min_row=7, max_col=18, values_only=True):
        municipality_code = str(values[0])
        if municipality_code not in CENTER_CODES:
            continue
        rows.append(
            {
                "municipality_code": municipality_code,
                "municipality_name_ja": str(values[1]),
                "population": float(values[5]),
                "households": float(values[17]),
                "population_date": "2022-10-01",
            }
        )
    dataframe = pd.DataFrame(rows)
    require_codes(dataframe, "municipality_code", CENTER_CODES, "population")
    return dataframe


def load_hospitals() -> pd.DataFrame:
    path = RAW / "mhlw_health" / "r5_chubu_ward_table.xlsx"
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    worksheet = workbook[workbook.sheetnames[0]]
    hospitals: dict[str, dict[str, Any]] = {}
    for values in worksheet.iter_rows(min_row=5, max_col=23, values_only=True):
        municipality_code = str(values[7])
        if municipality_code not in CENTER_CODES:
            continue
        hospital_id = str(values[0])
        record = hospitals.setdefault(
            hospital_id,
            {
                "hospital_id": hospital_id,
                "hospital_name_ja": str(values[1]),
                "municipality_code": municipality_code,
                "municipality_name_ja": str(values[8]),
                "general_beds": 0.0,
                "long_term_beds": 0.0,
                "ward_count": 0,
            },
        )
        record["general_beds"] += numeric(values[18])
        record["long_term_beds"] += numeric(values[22])
        record["ward_count"] += 1
    dataframe = pd.DataFrame(hospitals.values())
    dataframe["reported_beds"] = dataframe["general_beds"] + dataframe["long_term_beds"]
    require_codes(dataframe, "municipality_code", CENTER_CODES, "hospital")
    return dataframe.sort_values(["municipality_code", "hospital_name_ja"]).reset_index(drop=True)


def load_road_geometries() -> tuple[dict[str, tuple[float, float]], dict[str, dict[str, Any]]]:
    path = next((RAW / "mlit_roads_extracted" / "240125data").rglob("intercity_travel_time.geojson"))
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    coordinates: dict[str, tuple[float, float]] = {}
    route_features: dict[str, dict[str, Any]] = {}
    for feature in data["features"]:
        properties = feature.get("properties", {})
        name = str(properties.get("name", ""))
        geometry = feature.get("geometry", {})
        if geometry.get("type") == "Point":
            for code, municipality in MUNICIPALITIES.items():
                if municipality["name_ja"] in name:
                    lon, lat = geometry["coordinates"][:2]
                    coordinates[code] = (float(lat), float(lon))
        if geometry.get("type") == "LineString":
            for observation in CORRIDOR_OBSERVATIONS:
                if all(token in name for token in observation["feature_tokens"]):
                    route_features[observation["link_id"]] = feature
    require_codes(pd.DataFrame({"municipality_code": list(coordinates)}), "municipality_code", CENTER_CODES, "road point")
    missing_links = sorted({item["link_id"] for item in CORRIDOR_OBSERVATIONS} - set(route_features))
    if missing_links:
        raise ValueError(f"Missing road geometries for {missing_links}")
    return coordinates, route_features


def load_recovery_counts() -> Counter[str]:
    path = next((RAW / "mlit_roads_extracted" / "240112data").rglob("recovery_point.geojson"))
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    counts: Counter[str] = Counter()
    place_key = "地名（推定）"
    for feature in data["features"]:
        place = str(feature.get("properties", {}).get(place_key, ""))
        for code, municipality in MUNICIPALITIES.items():
            if municipality["name_ja"] in place:
                counts[code] += 1
                break
    return counts


def build_zones(population: pd.DataFrame) -> pd.DataFrame:
    dataframe = population[population["municipality_code"].isin(CORE_ZONE_CODES)].copy()
    dataframe["zone_id"] = dataframe["municipality_code"].map(lambda value: f"zone_{value}")
    dataframe["municipality_name_en"] = dataframe["municipality_code"].map(lambda value: MUNICIPALITIES[value]["name_en"])
    dataframe["full_destroyed_dwellings"] = dataframe["municipality_code"].map(
        lambda value: MUNICIPALITIES[value]["full_destroyed"]
    )
    dataframe["collapse_fraction"] = (dataframe["full_destroyed_dwellings"] / dataframe["households"]).clip(0.0, 1.0)
    dataframe["at_risk_population"] = dataframe["population"] * dataframe["collapse_fraction"]
    dataframe["renovation_cost"] = (dataframe["at_risk_population"] / 1000.0).clip(lower=1.0)
    dataframe["damage_date"] = "2024-10-01"
    dataframe["q_definition"] = "full_destroyed_dwellings / pre-event households"
    return dataframe.sort_values("municipality_code").reset_index(drop=True)


def build_centers(hospitals: pd.DataFrame, coordinates: dict[str, tuple[float, float]]) -> pd.DataFrame:
    beds = hospitals.groupby("municipality_code", as_index=False)["reported_beds"].sum()
    records: list[dict[str, Any]] = []
    for row in beds.itertuples(index=False):
        code = str(row.municipality_code)
        lat, lon = coordinates[code]
        share = OPERATIONAL_SHARES[code]
        records.append(
            {
                "center_id": f"center_{code}",
                "municipality_code": code,
                "municipality_name_en": MUNICIPALITIES[code]["name_en"],
                "municipality_name_ja": MUNICIPALITIES[code]["name_ja"],
                "latitude": lat,
                "longitude": lon,
                "reported_beds": float(row.reported_beds),
                "operational_share": share,
                "throughput_per_bed": THROUGHPUT_PER_BED,
                "existing_capacity": float(row.reported_beds) * share * THROUGHPUT_PER_BED,
                "capacity_unit_cost": 1.0,
                "capacity_status": "scenario-calibrated operational share; observed pre-event beds",
            }
        )
    dataframe = pd.DataFrame(records)
    require_codes(dataframe, "municipality_code", CENTER_CODES, "center")
    return dataframe.sort_values("municipality_code").reset_index(drop=True)


def build_corridors(route_features: dict[str, dict[str, Any]], recovery_counts: Counter[str]) -> pd.DataFrame:
    max_points = max(1, max(recovery_counts.values(), default=0))
    records: list[dict[str, Any]] = []
    for observation in CORRIDOR_OBSERVATIONS:
        feature = route_features[observation["link_id"]]
        length_km = line_length_km(feature["geometry"]["coordinates"])
        delay = observation["disrupted_minutes"] - observation["normal_minutes"]
        delay_ratio = delay / observation["disrupted_minutes"]
        endpoint_points = max(recovery_counts[observation["tail_code"]], recovery_counts[observation["head_code"]])
        verified_points = max(int(observation["recovery_points"]), int(endpoint_points))
        phi = min(0.62, max(0.08, 0.08 + 0.55 * delay_ratio + 0.12 * verified_points / max_points))
        records.append(
            {
                "link_id": observation["link_id"],
                "label": observation["label"],
                "tail_code": observation["tail_code"],
                "head_code": observation["head_code"],
                "normal_minutes": observation["normal_minutes"],
                "disrupted_minutes": observation["disrupted_minutes"],
                "failure_penalty_minutes": delay,
                "delay_ratio": delay_ratio,
                "initial_recovery_points": verified_points,
                "route_length_km": length_km,
                "baseline_failure_probability": phi,
                "retrofit_cost": 0.50 + length_km / 50.0,
                "disrupted_observation": observation["disrupted_observation"],
                "normal_observation": "2024-06-18 to 2024-06-21 directional mean",
                "phi_definition": "0.08 + 0.55*delay_ratio + 0.12*normalized_recovery_points; clipped [0.08,0.62]",
            }
        )
    return pd.DataFrame(records)


def summarize_snapshots() -> pd.DataFrame:
    records: list[dict[str, Any]] = []
    for snapshot_dir in sorted((RAW / "mlit_roads_extracted").iterdir()):
        if not snapshot_dir.is_dir():
            continue
        record: dict[str, Any] = {"snapshot": snapshot_dir.name.replace("data", "")}
        for filename, field in [
            ("emergency_restored_section.geojson", "restored_section_features"),
            ("recovery_point.geojson", "recovery_point_features"),
            ("intercity_travel_time.geojson", "intercity_features"),
            ("ETC2.0_speed_data.geojson", "speed_features"),
        ]:
            matches = list(snapshot_dir.rglob(filename))
            record[field] = feature_count(matches[0]) if matches else 0
        records.append(record)
    return pd.DataFrame(records)


def build_source_manifest() -> pd.DataFrame:
    records = [
        {
            "layer": "population_households",
            "status": "observed",
            "source": "Ishikawa Prefecture Statistics, 2022 municipality population and households",
            "url": "https://toukei.pref.ishikawa.lg.jp/library/2022.html",
            "local_path": str(RAW / "ishikawa_statistics" / "R4_population_machine_readable.xlsx"),
        },
        {
            "layer": "building_damage",
            "status": "observed",
            "source": "Ishikawa Prefecture, Noto damage report No. 162, 2024-10-01",
            "url": "https://www.pref.ishikawa.lg.jp/saigai/documents/higaihou_162_1001_1400.pdf",
            "local_path": str(RAW / "mlit_building_damage" / "ishikawa_damage_report_20241001.pdf"),
        },
        {
            "layer": "road_disruption_and_travel_time",
            "status": "observed",
            "source": "MLIT Noto road restoration map daily GeoJSON and travel-time tables",
            "url": "https://www.mlit.go.jp/road/r6noto/index2.html",
            "local_path": str(RAW / "mlit_roads_extracted"),
        },
        {
            "layer": "hospital_beds",
            "status": "observed",
            "source": "MHLW FY2023 Hospital Function Report",
            "url": "https://www.mhlw.go.jp/stf/seisakunitsuite/bunya/open_data_00016.html",
            "local_path": str(RAW / "mhlw_health" / "r5_chubu_ward_table.xlsx"),
        },
        {
            "layer": "hospital_operational_share",
            "status": "scenario-calibrated",
            "source": "Calibrated from documented Noto hospital water outages and emergency intake constraints",
            "url": "https://kouseikyoku.mhlw.go.jp/tokaihokuriku/000391260.pdf",
            "local_path": "",
        },
        {
            "layer": "retrofit_costs_and_budgets",
            "status": "scenario-calibrated",
            "source": "Route-length normalized costs and declared sector budget fractions",
            "url": "",
            "local_path": str(PREPARED / "noto_corridors.csv"),
        },
    ]
    dataframe = pd.DataFrame(records)
    dataframe["sha256"] = dataframe["local_path"].map(checksum_if_file)
    return dataframe


def build_coverage_table() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"model_variable": "P_rl", "status": "observed", "construction": "2022 municipality population"},
            {"model_variable": "q_rl", "status": "observed proxy", "construction": "full destroyed dwellings / pre-event households"},
            {"model_variable": "L", "status": "observed", "construction": "MLIT intercity road geometries"},
            {"model_variable": "tau", "status": "observed", "construction": "MLIT January disruption and June recovery travel times"},
            {"model_variable": "Phi_ij", "status": "constructed from observations", "construction": "delay ratio and restoration-point score"},
            {"model_variable": "w_k^0", "status": "mixed", "construction": "observed beds times declared throughput and operational shares"},
            {"model_variable": "C_rl,C_ij,lambda_k", "status": "scenario-calibrated", "construction": "normalized costs with sensitivity required"},
            {"model_variable": "B_Z,B_Y,B_X", "status": "scenario-calibrated", "construction": "declared fractions of sector cost/capacity totals"},
        ]
    )


def line_length_km(coordinates: list[list[float]]) -> float:
    return float(sum(haversine_km(first, second) for first, second in zip(coordinates, coordinates[1:])))


def haversine_km(first: list[float], second: list[float]) -> float:
    lon1, lat1 = first[:2]
    lon2, lat2 = second[:2]
    radius = 6371.0
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    value = math.sin(dphi / 2.0) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2.0) ** 2
    return 2.0 * radius * math.asin(math.sqrt(value))


def feature_count(path: Path) -> int:
    return len(json.loads(path.read_text(encoding="utf-8-sig")).get("features", []))


def checksum_if_file(value: str) -> str:
    path = Path(value) if value else None
    if path is None or not path.is_file():
        return ""
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def numeric(value: Any) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def require_codes(dataframe: pd.DataFrame, column: str, expected: list[str], layer: str) -> None:
    missing = sorted(set(expected) - set(dataframe[column].astype(str)))
    if missing:
        raise ValueError(f"Missing {layer} records for municipality codes {missing}")


def write_table(dataframe: pd.DataFrame, stem: str) -> None:
    atomic_write_dataframe(dataframe, PREPARED / f"{stem}.csv")


if __name__ == "__main__":
    main()
